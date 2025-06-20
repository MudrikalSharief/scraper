import requests
from bs4 import BeautifulSoup
import random
import time
import pandas as pd
from openpyxl import load_workbook

def header():
    headers_list = [
        {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"},
        {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"},
        {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0 Safari/605.1.15"},
        {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:90.0) Gecko/20100101 Firefox/90.0"},
        {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"},
    ]

    headers = random.choice(headers_list)
    return headers

def get_journal_links(url):
    links = []
    headers = header()
    headers.update({
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
    })
    time.sleep(random.uniform(1, 3))  # Mimic human behavior with a delay

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find the specific div containing the links
    link_holder = soup.find('div', class_='wp-block-column')
    
    # If the div is found, extract the links
    if link_holder:
        # Find all list items within the div
        for link in link_holder.find_all('li'):
            # get the anchor tag within the list item
            a_tag = link.find('a')
            # If the anchor tag exists and has an href attribute, extract the link
            if a_tag and 'href' in a_tag.attrs:
                href = a_tag['href']
                # Check if the href starts with 'http' to ensure it's a valid link
                if href.startswith('http'):
                    texts = (a_tag.text.strip() if a_tag else 'NONE')
                    # Append the link and its text to the list
                    links.append({
                        'Journal Name': texts,
                        'Website URL': href
                    })
    else:
        print("No links found in the specified div.")

    return links

def save_to_excel(links, Excel_file):
    # Convert to DataFrame
    df_new = pd.DataFrame(links)
    
    try:
        
        #load existing workbook
        book = load_workbook(Excel_file)
        sheet = book.active

        # Find the last row in the existing sheet
        last_row = sheet.max_row

        # Get column indices for "Journal Name" and "Website URL"
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        journal_col_idx = header_row.index("Journal Name") + 1  # +1 because openpyxl is 1-indexed
        url_col_idx = header_row.index("Website URL") + 1

        # Add new data starting from the first empty row
        for i, (_, row) in enumerate(df_new.iterrows(), 1):
            sheet.cell(row=last_row + i, column=journal_col_idx, value=row['Journal Name'])
            sheet.cell(row=last_row + i, column=url_col_idx, value=row['Website URL'])
        
        # Save the workbook
        book.save(Excel_file)
        print(f"Data has been added to {Excel_file}")
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        # Fall back to creating a new file
        df_new.to_excel(Excel_file, index=False)

url = 'https://beallslist.net/standalone-journals/'
excel_file = 'Predatory Features.xlsx'

Journals = get_journal_links(url)
saving_to_excel = save_to_excel(Journals, excel_file)

if __name__ == "__main__":
    print("Script executed successfully.")
    print(f"Total journals found: {len(Journals)}")
    print(f"Data saved to {excel_file}")